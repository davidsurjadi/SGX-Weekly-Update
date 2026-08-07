import openpyxl
import glob
import os
import re
import json
import csv
from datetime import datetime

RAW_DIR = "/mnt_raw_reports"  # placeholder, overridden by caller via env var below
RAW_DIR = os.environ.get("RAW_DIR", RAW_DIR)
OUT_DIR = os.environ.get("OUT_DIR", ".")

def find_header_row(ws, marker, col=1, max_row=40):
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v and marker in str(v):
            return r
    return None

def parse_weekly_top10(ws, report_week):
    rows = []
    inst_header = find_header_row(ws, "Top 10 Institution Net Buy")
    retail_header = find_header_row(ws, "Top 10 Retail Net Buy")

    def read_block(header_row, side):
        block = []
        for r in range(header_row + 1, header_row + 11):
            buy_name = ws.cell(row=r, column=1).value
            buy_code = ws.cell(row=r, column=2).value
            buy_val = ws.cell(row=r, column=3).value
            sell_name = ws.cell(row=r, column=4).value
            sell_code = ws.cell(row=r, column=5).value
            sell_val = ws.cell(row=r, column=6).value
            rank = r - header_row
            if buy_name and buy_code is not None:
                block.append({
                    "report_week": report_week, "side": side, "direction": "buy",
                    "rank": rank, "stock_name": str(buy_name).strip(),
                    "stock_code": str(buy_code).strip(), "value_sgd_m": buy_val
                })
            if sell_name and sell_code is not None:
                block.append({
                    "report_week": report_week, "side": side, "direction": "sell",
                    "rank": rank, "stock_name": str(sell_name).strip(),
                    "stock_code": str(sell_code).strip(), "value_sgd_m": sell_val
                })
        return block

    if inst_header:
        rows.extend(read_block(inst_header, "institutional"))
    if retail_header:
        rows.extend(read_block(retail_header, "retail"))
    return rows

SECTOR_COLS = ["Consumer Cyclicals", "Consumer Non-Cyclicals", "Energy/Oil & Gas",
               "Financial Services", "Health care", "Industrials",
               "Materials & Resources", "Real Estate (excl REITs)", "REITs",
               "Technology (Hardware/Software)", "Telcos", "Utilities"]

def parse_sector_sheet(ws, investor_type):
    rows = []
    # header row has 'Overall' in col A around row 1-2, data starts after header row containing a datetime in col B
    header_row = None
    for r in range(1, 6):
        v = ws.cell(row=r, column=2).value
        if v and ("Investor" in str(v) or "investor" in str(v)):
            header_row = r
            break
    if header_row is None:
        return rows
    r = header_row + 1
    while True:
        overall = ws.cell(row=r, column=1).value
        date_val = ws.cell(row=r, column=2).value
        if date_val is None or not hasattr(date_val, "isoformat"):
            break
        rec = {"week_date": date_val.date().isoformat(), "investor_type": investor_type, "overall_sgd_m": overall}
        for i, sector in enumerate(SECTOR_COLS):
            rec[sector] = ws.cell(row=r, column=3 + i).value
        rows.append(rec)
        r += 1
    return rows

def main():
    files = sorted(glob.glob(os.path.join(RAW_DIR, "*.xlsx")))
    all_top10 = []
    sector_by_key = {}  # (investor_type, week_date) -> record

    for fp in files:
        fname = os.path.basename(fp)
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
        except Exception as e:
            print(f"FAILED to open {fname}: {e}")
            continue

        # report week from the 'Weekly Top 10' sheet cell A1, e.g. "Week of 29 June 2026"
        report_week = None
        if "Weekly Top 10" in wb.sheetnames:
            ws = wb["Weekly Top 10"]
            a1 = ws.cell(row=1, column=1).value
            if a1:
                m = re.search(r"Week of (.+)", str(a1))
                if m:
                    report_week = m.group(1).strip()
            top10_rows = parse_weekly_top10(ws, report_week or fname)
            all_top10.extend(top10_rows)
        else:
            print(f"NO 'Weekly Top 10' sheet in {fname}")

        if "Institutional" in wb.sheetnames:
            for rec in parse_sector_sheet(wb["Institutional"], "institutional"):
                sector_by_key[(rec["investor_type"], rec["week_date"])] = rec
        if "Retail" in wb.sheetnames:
            for rec in parse_sector_sheet(wb["Retail"], "retail"):
                sector_by_key[(rec["investor_type"], rec["week_date"])] = rec

    # normalize report_week to an ISO date (Monday date of that week) for sorting
    def parse_week_label(label):
        # e.g. "29 June 2026" or "6 October 2025"
        label = label.replace("Sept ", "Sep ")
        for fmt in ("%d %B %Y", "%d %b %Y"):
            try:
                return datetime.strptime(label, fmt).date().isoformat()
            except ValueError:
                continue
        return None

    for row in all_top10:
        row["week_start"] = parse_week_label(row["report_week"]) if row["report_week"] else None

    # dedupe top10 rows (in case of any duplicate file coverage) by (week_start, side, direction, rank)
    dedup = {}
    for row in all_top10:
        key = (row["week_start"], row["side"], row["direction"], row["rank"])
        dedup[key] = row
    all_top10 = list(dedup.values())
    all_top10.sort(key=lambda x: (x["week_start"] or "", x["side"], x["direction"], x["rank"]))

    sector_rows = sorted(sector_by_key.values(), key=lambda x: (x["week_date"], x["investor_type"]))

    os.makedirs(OUT_DIR, exist_ok=True)

    with open(os.path.join(OUT_DIR, "weekly_top10.csv"), "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["week_start", "report_week", "side", "direction", "rank", "stock_name", "stock_code", "value_sgd_m"])
        w.writeheader()
        for row in all_top10:
            w.writerow({k: row.get(k) for k in w.fieldnames})

    with open(os.path.join(OUT_DIR, "sector_flow.csv"), "w", newline="") as f:
        fieldnames = ["week_date", "investor_type", "overall_sgd_m"] + SECTOR_COLS
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in sector_rows:
            w.writerow(row)

    print(f"Files processed: {len(files)}")
    print(f"weekly_top10 rows: {len(all_top10)}")
    print(f"distinct weeks in top10: {len(set(r['week_start'] for r in all_top10))}")
    print(f"sector_flow rows: {len(sector_rows)}")
    print(f"distinct tickers: {len(set(r['stock_code'] for r in all_top10))}")

if __name__ == "__main__":
    main()
